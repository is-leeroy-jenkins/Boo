'''
  ******************************************************************************************
      Assembly:                Boo
      Filename:                fsys.py
      Author:                  Terry D. Eppler
      Created:                 05-31-2022

      Last Modified By:        Terry D. Eppler
      Last Modified On:        05-01-2025
  ******************************************************************************************
  <copyright file="fsys.py" company="Terry D. Eppler">

	     Boo is a df analysis tool integrating various Generative AI, Text-Processing, and
	     Machine-Learning algorithms for federal analysts.
	     Copyright ©  2022  Terry Eppler

     Permission is hereby granted, free of charge, to any person obtaining a copy
     of this software and associated documentation files (the “Software”),
     to deal in the Software without restriction,
     including without limitation the rights to use,
     copy, modify, merge, publish, distribute, sublicense,
     and/or sell copies of the Software,
     and to permit persons to whom the Software is furnished to do so,
     subject to the following conditions:

     The above copyright notice and this permission notice shall be included in all
     copies or substantial portions of the Software.

     THE SOFTWARE IS PROVIDED “AS IS”, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
     INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
     FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT.
     IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
     DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
     ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
     DEALINGS IN THE SOFTWARE.

     You can contact me at:  terryeppler@gmail.com or eppler.terry@epa.gov

  </copyright>
  <summary>
    fsys.py
  </summary>
  ******************************************************************************************
  '''
from __future__ import annotations

import os
import zipfile as zp
from openpyxl import Workbook
import shutil as sh
from boogr import Error, ErrorDialog
from typing import Any, List, Tuple, Optional


class Path( ):
	'''

	Constructor:

		Path( path: str )

	Purpose:

		Class representing a url

	'''
	
	
	def __init__( self, filepath: str ):
		self.input = filepath
		self.name = os.path.split( filepath )[ 1 ]
		self.current_directory = os.getcwd( )
		self.file_extension = os.path.splitext( filepath )[ 1 ]
		self.parent_directory = os.path.split( filepath )[ 0 ]
		self.template = r'resources\templates\report\Report.xlsx'
		self.path_separator = os.path.sep
		self.extension_separator = os.extsep
		self.drive_separator = ':\\'
		self.drive = os.path.splitdrive( filepath )[ 0 ]
	
	
	def __str__( self ) -> str | None:
		if self.input is not None:
			return str( self.input )
	
	
	def __dir__( self ) -> List[ str ]:
		'''

		Returns a get_list[ str ] of member names.

		'''
		return [ 'path', 'name', 'current_directory', 'extension',
		         'parent_directory', 'path_separator', 'drive',
		         'drive_separator', 'extension_separator', 'internal_path',
		         'exists', 'is_folder', 'is_file', 'is_absolute',
		         'is_relative', 'verify', 'join', 'copy_tree' ]
	
	
	def exists( self ) -> bool | None:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if not os.path.exists( self.input ):
				_msg = "The object 'self' does not exist or is None!"
				raise Exception( _msg )
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'exists( self ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def is_folder( self ) -> bool | None:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if self.input is None:
				_msg = "The 'path' is None!"
				raise Exception( _msg )
			if not os.path.isdir( self.input ):
				return False
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'is_folder( self ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def is_file( self ) -> bool | None:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if self.input is None:
				_msg = "The 'path' is None!"
				raise Exception( _msg )
			elif not os.path.isfile( self.input ):
				return False
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'is_file( self ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def is_absolute( self ) -> bool | None:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if not os.path.isabs( self.input ):
				_msg = "The object 'self' is not a file!"
				raise Exception( _msg )
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'is_absolute( self ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def is_relative( self ) -> bool | None:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if not os.path.isabs( self.input ):
				_msg = "The object 'self' is not a file!"
				raise Exception( _msg )
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'is_relative( self ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def verify( self, other: str ) -> bool | None:
		'''
		
		
			Purpose:
			Verifies the file 'other' exists
	
			Parameters:
	
			Returns:
			
			
		'''
		try:
			if not os.path.exists( other ):
				_msg = "The object 'self' is not a url!"
				raise Exception( _msg )
			else:
				return True
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'verify( self, other: str ) -> bool'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def join( self, first: str, second: str ) -> str | None:
		'''

			Purpose:
			Joins two paths into one.
	
			Parameters: 'first: str' representing the first url
			that is joined to the second url 'second: str'.
	
			Returns: a single path representing a url

		'''
		try:
			if not os.path.exists( first ) or not os.path.exists( second ):
				_msg = "The arguement 'first' or 'second' is null!"
				raise Exception( _msg )
			else:
				return os.path.join( first, second )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'join( self, first: str, second: str ) -> str'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def copy_tree( self, destination: str ) -> None:
		'''


			Purpose:
			    Copies directory tree to 'destination'
	
			Parameters:
	
			    destination: str
	
			Returns:
	
			Void

		'''
		try:
			if destination is None:
				_msg = "The argument 'destination' is None"
				raise Exception( _msg )
			else:
				sh.copytree( self.input, destination )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'copy_tree( self, destination: str ) -> None'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def create_link( self, path: str, name: str ):
		'''

			Purpose:
			Creates a symbolic link of 'url' given the name 'name'
	
			Parameters:
			url: str, name: str
	
			Returns:
			str

		'''
		try:
			if path is None:
				_msg = "The argument 'url' is None"
				raise Exception( _msg )
			elif name is None:
				_msg = "The argument 'name' is None"
				raise Exception( _msg )
			else:
				os.link( path, name )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'create_link( self, url: str, name: str )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def read_link( self, path: str ):
		'''

			Purpose:
	
				reads a symbolic link of 'url'
	
			Parameters:
	
				path: str, name: str
	
			Returns:
	
				str

		'''
		try:
			if path is None:
				_msg = "The argument 'url' is None"
				raise Exception( _msg )
			else:
				os.readlink( path )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Path'
			_exc.method = 'read_link( self, url: str )'
			_err = ErrorDialog( _exc )
			_err.show( )


class File( Path ):
	'''
	
		Constructor:
	
			File( url: str )
	
		Purpose:
	
			Class providing file information

	 '''
	
	
	def __init__( self, path: str ):
		super( ).__init__( path )
		self.input = super( ).input
		self.absolute_path = os.path.abspath( path )
		self.relative_path = os.path.relpath( path )
		self.name = os.path.basename( path )
		self.size = os.path.getsize( path )
		self.extension = super( ).file_extension
		self.created = os.path.getctime( path )
		self.accessed = os.path.getatime( path )
		self.modified = os.path.getmtime( path )
	
	
	def __str__( self ) -> str | None:
		if self.relative_path is not None:
			return self.relative_path
	
	
	def __dir__( self ) -> List[ str ]:
		'''

		Returns a get_list[ str ] of member names.

		'''
		return [ 'name', 'current_directory', 'extension',
		         'parent_directory', 'path_separator', 'drive',
		         'drive_separator', 'extension_separator', 'internal_path',
		         'exists', 'is_folder', 'is_file', 'is_absolute',
		         'is_relative', 'verify', 'join', 'copy_tree',
		         'absolute_path', 'relative_path', 'url',
		         'name', 'size', 'extension', 'created',
		         'accessed', 'modified', 'rename', 'move',
		         'generate_text', 'delete', 'get_lines', 'readlines', 'readall',
		         'iterlines', 'writelines', 'writeall' ]
	
	
	def rename( self, other: str ) -> str:
		'''

			Purpose:
	
			Parameters:
	
			Returns:

		'''
		try:
			if other is None:
				_msg = " The argument 'other' has not been specified!"
				raise Exception( _msg )
			else:
				_src = os.path.abspath( self.path )
				_dest = os.path.join( self.directory, other )
				os.rename( _src, _dest )
				return _dest
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'rename( self, other: str ) -> str'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def move( self, destination: str ) -> str:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			_msg = " The 'url' is not a file or the argument " \
			       "'destination' has not been specified!"
			if os.path.isfile( self.input ) == False or destination is None:
				raise Exception( _msg )
			else:
				return os.path.join( self.name, destination )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'move( self, destination ) -> str'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def create( self, other: str, lines: List[ str ]):
		'''
		
			Purpose:
			creates a file 'other' and writes 'words' to it
	
			Parameters: other: str
	
			Returns: void
			
		'''
		try:
			if other is None:
				raise Exception( 'The argument "other" has not been specified!' )
			elif lines is None:
				raise Exception( 'The argument "words" has not been specified!' )
			else:
				_file = open( other, 'r+' )
				if len( lines ) > 0:
					for line in lines:
						_file.write( line )
					_file.flush( )
					_file.close( )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'generate_text( self, other, words = None )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def delete( self, other: str ):
		'''
		
			Purpose:
			deletes a file provided its url 'other'
	
			Parameters:
			other: str
	
			Returns:
			void
			
		'''
		try:
			if other is None:
				raise Exception( "The argument 'other' does not exist!" )
			else:
				os.remove( other )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'delete( self, other )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def get_lines( self ) -> List[ str ]:
		'''
		
			Purpose:
			reads 'self.path' appending each line to a get_list
	
			Parameters: void
	
			Returns: get_list[ str ]
			
		'''
		_lines = list( )
		try:
			if os.path.isfile( self.input ) == False:
				_msg = "The 'url' is not a file!"
				raise Exception( _msg )
			else:
				file = open( self.input )
				for i in file.readlines( ):
					_lines.append( i )
				return _lines
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'get_lines( self ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def iterlines( self ) -> iter:
		'''
		
			Purpose:
			iterates words of 'self.path'
	
			Parameters:
			void
	
			Returns:
			Generator
			
		'''
		try:
			if os.path.isfile( self.input ) == False:
				_msg = "The 'url' is not a file!"
				raise Exception( _msg )
			else:
				file = open( self.input )
				for i in file.readlines( ):
					yield i
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'iterlines( self ) -> iter'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def readlines( self ) -> List[ str ]:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			_contents = list( )
			if self.input is None  :
				_msg = "The 'path' is not a file!"
				raise Exception( _msg )
			else:
				_file = open( self.input )
				for i in _file.readlines( ):
					_contents.append( i )
				_file.close( )
				return _contents
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'readlines( self ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def readall( self ) -> str:
		'''
		
			Purpose:
	
			Parameters:
	
			Returns:
			
		'''
		try:
			_contents = ''
			if not os.path.isfile( self.input ):
				_msg = "The 'path' is not a file!"
				raise Exception( _msg )
			else:
				_file = open( self.input )
				_contents = _file.read( )
				_file.close( )
				return _contents
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'readall( self ) -> str'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def writelines( self, lines: List[ str ] ):
		'''
		
			Purpose:
			writes words in 'words' to file 'self.path'
	
			Parameters: get_list[ str ]
	
			Returns: void
			
		'''
		try:
			if lines is None or os.path.isfile( self.input ) == False:
				_msg = "The 'words' is None or 'url' is not a file!"
				raise Exception( _msg )
			else:
				_path = os.path.relpath( self.input )
				_contents = open( _path, 'a' )
				for line in lines:
					_contents.write( line )
				_contents.flush( )
				_contents.close( )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'writelines( self, words: get_list[ str ] ):'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def writeall( self, other: str ) -> List[ str ]:
		'''

			Purpose:
			writes all documents in 'other' to file 'self.path'
	
			Parameters: str
	
			Returns: void

		'''
		try:
			_contents = [ ]
			_lines = [ ]
			if other is None or os.path.isfile( self.input ) == False:
				_msg = "The argument 'other' has not been specified " \
				       "or the 'url' is not a file!"
				raise Exception( _msg )
			else:
				_path = os.path.relpath( self.input )
				_contents = open( _path, 'a' )
				_lines = open( other )
				for line in _lines.readlines( ):
					_contents.write( line )
				_contents.flush( )
				_lines.close( )
				return _contents
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'File'
			_exc.method = 'writeall( self, other: str ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )


class Folder( Path ):
	'''
	
		Constructor:
	
		    Folder( path: str )
	
	    Purpose:
	
	        Class providing file directory information

	'''
	def __init__( self, filepath: str ):
		super( ).__init__( filepath )
		self.input = super( ).input
		self.name = super( ).name
		self.size = os.path.getsize( filepath )
		self.absolute_path = os.path.abspath( filepath )
		self.relative_path = f'{os.getcwd( )}\\{os.path.basename( filepath )}'
	
	
	def __str__( self ) -> str:
		if self.input is not None:
			return self.input
	
	
	def __dir__( self ) -> List[ str ]:
		'''

		Returns a get_list[ str ] of member names.

		'''
		return [ 'name', 'current_directory', 'extension',
		         'parent_directory', 'path_separator', 'drive',
		         'drive_separator', 'extension_separator', 'internal_path',
		         'exists', 'is_folder', 'is_file', 'is_absolute',
		         'is_relative', 'verify', 'join', 'copy_tree',
		         'absolute_path', 'relative_path', 'url',
		         'name', 'size', 'get_files', 'get_subfiles',
		         'get_subfolders', 'rename', 'move', 'generate_text',
		         'delete', 'iterate' ]
	
	
	def get_files( self ) -> List[ str ]:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			_names = [ ]
			_msg = "The object 'self' is not a directory!"
			if not os.path.isdir( self.input ):
				raise Exception( _msg )
			else:
				for i in os.listdir( self.input ):
					_path = os.path.join( self.absolute_path, i )
					if os.path.isfile( _path ):
						_names.append( _path )
				return _names
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'get_files( self ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def get_subfiles( self ) -> List[ str ]:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			_names = [ ]
			_msg = "The path url 'path' is not a directory!"
			if not os.path.isdir( self.input ):
				raise Exception( _msg )
			else:
				for i in os.walk( self.input ):
					_dirpath = i[ 0 ]
					if len( i[ 1 ] ) > 0:
						for name in i[ 1 ]:
							path = os.path.join( _dirpath, name )
							_names.append( path )
					return _names
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'get_subfiles( self ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def get_subfolders( self ) -> List[ str ]:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			_names = [ ]
			if not os.path.isdir( self.input ):
				_msg = "The object 'self' is not a directory!"
				raise Exception( _msg )
			else:
				for i in os.walk( self.absolute_path ):
					if len( i[ 1 ] ) > 0:
						for file in i[ 1 ]:
							_path = os.path.join( self.absolute_path, file )
							if os.path.isdir( _path ):
								_names.append( _path )
				return _names
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'get_subfolders( self ) -> get_list[ str ]'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def rename( self, name: str ):
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if name is None:
				_msg = "The argument 'name' has not been specified!"
				raise Exception( _msg )
			else:
				return os.rename( self.name, name )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'rename( self, name: str )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def move( self, destination: str ):
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		try:
			if destination is None:
				_msg = "The 'destination' has not been specified!"
				raise Exception( _msg )
			else:
				return os.path.join( self.name, destination )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'move( self, destination: str )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def create( self, other: str ):
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		
		try:
			if other is None:
				_msg = "The argument 'other' has not been specified!"
				raise Exception( _msg )
			else:
				os.mkdir( other )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'generate_text( self, other: str )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def delete( self, other: str ):
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		
		try:
			if other is None:
				_msg = "The argument 'other' doesn't exist or has not been specified!"
				raise Exception( _msg )
			else:
				os.rmdir( other )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'delete( self, other: str )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def iterate( self ) -> iter:
		'''
			Purpose:
	
			Parameters:
	
			Returns:
		'''
		
		try:
			if not os.path.isdir( self.input ):
				_msg = " 'self' is not a directory!"
				raise Exception( _msg )
			else:
				for i in os.walk( self.input ):
					yield i
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Folder'
			_exc.method = 'iterate( self ) -> iter'
			_err = ErrorDialog( _exc )
			_err.show( )


class Message( ):
	'''

		Constructor:
	
		    GptMessage( sender: str, receiver: str,
		              body: str, subject: str, copy: str = '')
	
		Purpose:
	
		    Class providing email behavior

	'''
	def __init__( self, sender: str, receiver: List[ str ], body: List[ str ],
	              subject: str, copy: List[ str ]=None ):
		self.sender = sender
		self.receiver = receiver
		self.body = body
		self.others = copy
		self.subject = subject
	
	
	def __str__( self ) -> List[ str ] | None:
		if self.body is not None:
			return self.body
	
	
	def __dir__( self ) -> List[ str ]:
		'''

			Returns a get_list[ str ] of member names.

		'''
		return [ 'sender', 'receiver', 'subject',
		         'body', 'copy' ]


class Email( Message ):
	'''

		Constructor:
	
			Emai( sender: str, receiver: str, body: str,
					  subject: str, copy: str = '' )
	
		Purpose:
	
			Class providing email behavior

	'''
	def __init__( self, sender: str, receiver: List[ str ], body: List[ str ],
	              subject: str, copy: List[ str ]=None ):
		super( ).__init__( sender, receiver, body, subject, copy )
		self.sender = super( ).sender
		self.receiver = super( ).receiver
		self.body = super( ).body
		self.others = super( ).others
		self.subject = super( ).subject
	
	
	def __dir__( self ) -> List[ str ]:
		'''

			Returns a get_list[ str ] of member names.

		'''
		return [ 'sender', 'receiver', 'subject',
		         'body', 'copy' ]


class Excel( ):
	'''

		Constructor:
	
			Excel( url: str )
	
		Purpose:
	
			Class provides the spreadsheet for reports

	'''
	def __init__( self, path: str ):
		self.internal_path = r'resources/templates/report/Excel.xlsx'
		self.external_path = path
		self.name = os.path.split( path )[ 1 ]
		self.title = self.name.split( '.' )[ 0 ]
		self.workbook = Workbook( )
		self.worksheet = Workbook( ).create_sheet( self.title, 0 )
	
	
	def __str__( self ) -> str | None:
		if self.external_path is not None:
			return self.external_path
	
	
	def __dir__( self ) -> List[ str ]:
		'''

			Returns a get_list[ str ] of member names.

		'''
		return [ 'internal_path', 'external_path', 'name',
		         'title', 'workbook', 'worksheet', 'save' ]
	
	
	def save( self ):
		'''

			Purpose:
	
			Parameters:
	
			Returns:

		'''
		try:
			self.workbook.save( self.external_path )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'Excel'
			_exc.method = 'save( self )'
			_err = ErrorDialog( _exc )
			_err.show( )


class ExcelReport( Excel ):
	'''
	
		Constructor:
	
			ExcelReport( url: str, rows: int = 46, cols: int = 12 ).
	
		Purpose:
	
			Class providing spreadsheet for reports

	'''
	def __init__( self, path: str, rows: int=46, cols: int=12 ):
		super( ).__init__( path )
		self.internal = super( ).internal_path
		self.path = super( ).external_path
		self.name = super( ).name
		self.rows = rows
		self.columns = cols
		self.dimensions = (rows, cols)
		self.title = super( ).title
		self.workbook = super( ).workbook
		self.worksheet = super( ).worksheet
	
	
	def __dir__( self ) -> List[ str ]:
		'''

		Returns a get_list[ str ] of member names.

		'''
		return [ 'rows', 'columns', 'dimensions' ]


class ZipFile( ):
	'''
	
		Constructor:
	
			ZipFile( url: str )
	
		Purpose:
	
			Class defines object providing zip file functionality

	'''
	def __init__( self, path: str ):
		self.input = path
		self.zip_extension = '.zip'
		self.file_path = path
		self.file_extension = os.path.splitext( path )[ 1 ]
		self.zip_path = self.file_path.replace( self.file_extension, self.zip_extension )
		self.file_name = os.path.basename( path )
		self.zip_name = self.file_name + self.zip_extension
	
	
	def __str__( self ):
		if self.file_path is not None:
			return self.file_path
	
	
	def __dir__( self ) -> List[ str ]:
		'''

			Returns a get_list[ str ] of member names.

		'''
		return [ 'file_name', 'authority_filepath', 'file_extenstion',
		         'zip_name', 'zip_path', 'zip_extension',
		         'generate_text', 'unzip' ]
	
	
	def create( self ):
		'''
	
			Purpose: creates a zipfile from a given url
	
			Parameters: void
	
			Returns: void

		'''
		try:
			if self.file_path is not None:
				(zp.ZipFile( self.zip_path, 'w' )
				 .write( self.file_path, self.file_name ) )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'ZipFile'
			_exc.method = 'generate_text( self )'
			_err = ErrorDialog( _exc )
			_err.show( )
	
	
	def unzip( self ):
		'''
		
				Purpose: Un-zips the file
		
				Parameters: void
		
				Returns: void
				
		'''
		try:
			if os.path.exists( self.zip_path ):
				_file = zp.ZipFile( self.zip_path )
				_file.extractall( self.zip_path )
		except Exception as e:
			_exc = Error( e )
			_exc.module = 'fsys'
			_exc.cause = 'ZipFile'
			_exc.method = 'unzip( self )'
			_err = ErrorDialog( _exc )
			_err.show( )
